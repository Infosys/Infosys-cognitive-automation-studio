'''
Copyright 2020 Infosys Ltd.
Use of this source code is governed by Apache 2.0 license that can be found in the LICENSE file or at 
https://opensource.org/licenses/Apache-2.0 .
'''
import openpyxl as xl
from abstract_bot import Bot #importing the Botclass from the abstract_bot.py
#Bot to copy and paste filtered data into new sheet in same excel
class CopyFiltereddataInExcel(Bot):
    
    def bot_init(self):
        pass
    
    def execute(self,executionContext):
        try:
            filePath = executionContext["filePath"] #Path of the excel file
            sheetName = executionContext["sheetName"] #Sheetname to be filtered
            sheetNameCreate = executionContext["sheetNameCreate"] #Name of the new sheet where the filtered data will be pasted
            wb = xl.load_workbook(filePath)
            ws1 = wb[sheetName]
            ws2 = wb.create_sheet(sheetNameCreate)
            for row in ws1:
                if ws1.row_dimensions[row[0].row].hidden == False:
                    for cell in row:
                        ws2[cell.coordinate].value = cell.value
            indexRow = []
            for i in range(1, ws2.max_row):
                if ws2.cell(i, 1).value is None:
                    indexRow.append(i)
            for rowDel in range(len(indexRow)):
                ws2.delete_rows(idx=indexRow[rowDel], amount=1)
                indexRow = list(map(lambda k: k - 1, indexRow))
            wb.save(filePath)
            wb.close()
            return {'Result': 'Success'}
        except Exception as e:
            return {'Exception' : str(e)}
        
if __name__ == "__main__":
    context = {}
    bot_obj = CopyFiltereddataInExcel()
    context = {'filePath':'','sheetName':'','sheetNameCreate':''}
    bot_obj.bot_init()
    output = bot_obj.execute(context)
    print(output)