'''
Copyright 2020 Infosys Ltd.
Use of this source code is governed by Apache 2.0 license that can be found in the LICENSE file or at 
https://opensource.org/licenses/Apache-2.0 .
'''
import win32com.client
import openpyxl 
from difflib import SequenceMatcher
from rake_nltk import Rake
from nltk.corpus import stopwords 
from abstract_bot import Bot

class GenerateResponseToEmail(Bot):
    
    def bot_init(self):
        pass
    
    def execute(self,executionContext):
        try:
            excelPath=executionContext['excelPath']
            
            if excelPath=='':
                return{"Missing Argument": "excelPath"}
            
            #Loading the Solution_Repository excel file
            wbObj = openpyxl.load_workbook(excelPath)
            sheetObj = wbObj.active 
            mRow = sheetObj.max_row
            
            #Triggering the Outlook application
            outlook =win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            inbox = outlook.GetDefaultFolder(6) # 6 is used for the index of the folder
            msgBox = inbox.Items
            
            r=Rake()
           
            
            for message in msgBox:
                subKeywords=[]
                bodyKeywords=[]
                if message.UnRead==True:
                    probSubject=message.subject
                    a=r.extract_keywords_from_text(probSubject)
                    b=r.get_ranked_phrases_with_scores()
                        
                    for key in b:
                        subKeywords.append(key[1])  
                    probSubKeys=subKeywords[0:4]
                    
                        
                    probBody=message.Body
                    a1=r.extract_keywords_from_text(probBody)
                    b1=r.get_ranked_phrases_with_scores()
                        
                    for key1 in b1:
                        if key1[0]>1:
                            bodyKeywords.append(key1[1])  
                    probBodyKeys=bodyKeywords[0:4]
                    
                    
                    probBodyStr=''
                    probSubStr=''
                   
                    for p in probBodyKeys:
                        probBodyStr+=str(p)+'-'
                   
                    
                    for q in probSubKeys:
                        probSubStr+=str(q)+'-'
                    
                    #looping through the Solution_Repository excel file
					#Solution_Repository excel file contains 4 columns viz. Subject, Body, Subject_Keys, Body_Keys
					
                    for i in range(1, mRow + 1): 
                        cellObj1 = sheetObj.cell(row = i, column = 1)
                        cellObj2 = sheetObj.cell(row=i, column=2)
                        cellObj3 = sheetObj.cell(row=i, column=3)
                        cellObj4 = sheetObj.cell(row=i, column=4)
                        
                        colSubValue=cellObj1.value
                        colBodyValue=cellObj2.value
                        colSubKey=cellObj3.value
                        colBodyKey=cellObj4.value
                        
                        
                        colBodyValue1=colBodyValue.replace('\r','')
                        colBodyValue2=colBodyValue1.replace('_x000D_','')
                                            
                                
                        ratio1=SequenceMatcher(None,probSubStr,colSubKey).ratio()
                        ratio2=SequenceMatcher(None,probBodyStr,colSubKey).ratio()
                        ratio3=SequenceMatcher(None,probBodyStr,colBodyKey).ratio()
                        ratio4=SequenceMatcher(None,probSubStr,colBodyKey).ratio()
                        
                        
                        if ratio1>0.4 or ratio2>0.4 or ratio3>0.4 or ratio4>0.4:
                            reply = message.Reply() 
                            reply.Body = colBodyValue2 
                            reply.Send()
                            message.Unread=False
                            #print("reply sent")
                            break
                        
            return{'Result':'Successfully replied to the mails from the Solution_Repository'}
        except Exception as e:
            return{'Exception':str(e)}
        
if __name__ == "__main__":
    context = {}
    bot_obj = GenerateResponseToEmail()
    #provide the path as "C:\\Users\\vikas.singh09\\Desktop\\Solution_Repository.xlsx"
    context={"excelPath": ""} #provide the path to Solution_Repository excel file
    bot_obj.bot_init()
    output = bot_obj.execute(context)
    print(output)
    
    