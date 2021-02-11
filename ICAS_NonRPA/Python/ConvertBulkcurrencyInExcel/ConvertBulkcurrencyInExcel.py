'''
Copyright 2020 Infosys Ltd.
Use of this source code is governed by Apache 2.0 license that can be found in the LICENSE file or at 
https://opensource.org/licenses/Apache-2.0 .
'''
# importing openpyxl module 
import openpyxl 
from forex_python.converter import CurrencyRates
from datetime import date

#import xlwt 
from xlwt import Workbook 
from abstract_bot import Bot
class ConvertBulkcurrencyInExcel(Bot):
    
    def bot_init(self):
        pass
    
    def execute(self,executionContext):
        try:
            sourcePath=executionContext['sourcePath']
            destinationPath=executionContext['destinationPath']
            fileName=executionContext['fileName']    
            today = date.today()
            # workbook object is created 
            wbObj = openpyxl.load_workbook(sourcePath) 
            # Workbook is created 
            wb = Workbook()   
            # add_sheet is used to create sheet.
            sheet1 = wb.add_sheet('Sheet 1') 

            sheetObj = wbObj.active 
            maxCol = sheetObj.max_column
            maxRow=sheetObj.max_row 

            # Loop will print all columns name 
            for i in range(1, maxRow + 1): 
                for j in range(1,maxCol + 1):
                    cellObj = sheetObj.cell(row = i, column = j)
                    if j==1:
                        sourceCurrency=cellObj.value
                        sheet1.write(i-1, j-1, sourceCurrency)
            
                    elif j==2:
                        sourceCurrencyAmount=cellObj.value
                        sheet1.write(i-1, j-1, sourceCurrencyAmount)
                    elif j==3:
                        destinationCurrency=cellObj.value
                        sheet1.write(i-1, j-1, destinationCurrency)
                    elif j==4:
                        if i==1:
                            destinationCurrencyAmount=cellObj.value
                            sheet1.write(i-1,j-1,destinationCurrencyAmount)
                        else:
                            c = CurrencyRates()        
                            cellObj.value=c.convert(sourceCurrency,destinationCurrency,sourceCurrencyAmount,today)
                            destinationCurrencyAmount=round(cellObj.value,3)
                            
                            sheet1.write(i-1,j-1,destinationCurrencyAmount)
                            
                        
            wbObj.save(destinationPath+fileName)
            return {'Status':'Conversion Completed and output is saved with the given name in given path'}
        except Exception as e:
            return {'Exception' : str(e)}
        
if __name__ == "__main__":
    context = {}
    bot_obj = ConvertBulkcurrencyInExcel()
    context={'sourcePath':'','destinationPath':'','fileName':''}
    
    #context = {'sourcePath':'C:\\Users\\anuj.gupta03\\Downloads\\demo11.xlsx','destinationPath':'C:\\Users\\anuj.gupta03\\Downloads\\','fileName':'example111.xlsx'}
               

   
    output = bot_obj.execute(context)
    print(output) 