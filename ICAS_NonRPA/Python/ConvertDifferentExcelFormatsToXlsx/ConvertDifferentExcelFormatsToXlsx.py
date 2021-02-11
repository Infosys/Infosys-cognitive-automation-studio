'''
Copyright 2020 Infosys Ltd.
Use of this source code is governed by Apache 2.0 license that can be found in the LICENSE file or at 
https://opensource.org/licenses/Apache-2.0 .
'''
# -*- coding: utf-8 -*-
"""
Created on Wed May 13 10:50:47 2020

@author: vaddi.kumar01
"""
#import pyxlsb
import pandas as pd
import xlrd
from openpyxl.workbook import Workbook
from abstract_bot import Bot

class ConvertDifferentExcelFormatsToXlsx(Bot):
    def execute(self, excecuteContext):
        try:
            inputFileName = excecuteContext['inputFileName']
            outputFileName = excecuteContext['outputFileName']

            #csv to xlsx
            def csv_to_xlsx(src_file_path, dst_file_path):
                read_file = pd.read_csv (src_file_path)
                read_file.to_excel (dst_file_path, index = None, header=True)
            
            #xlsb to xlsx
            def xlsb_to_xlsx(src_file_path, dst_file_path):
                read_file= pd.read_excel(inputFileName, engine='pyxlsb', sheet_name= None)               
                writer = pd.ExcelWriter(dst_file_path, engine='xlsxwriter')
                for sheet_name in read_file.keys():
                    read_file[sheet_name].to_excel(writer, sheet_name=sheet_name, index= False)
                writer.save()
            
            def xls_to_xlsx(src_file_path, dst_file_path):
                book_xls = xlrd.open_workbook(src_file_path)
                book_xlsx = Workbook()
            
                sheet_names = book_xls.sheet_names()
                for sheet_index, sheet_name in enumerate(sheet_names):
                    sheet_xls = book_xls.sheet_by_name(sheet_name)
                    if sheet_index == 0:
                        sheet_xlsx = book_xlsx.active
                        sheet_xlsx.title = sheet_name
                    else:
                        sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)
                            
                    for row in range(0, sheet_xls.nrows):
                        for col in range(0, sheet_xls.ncols):
                            sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
                book_xlsx.save(dst_file_path)
            
            
            if inputFileName.endswith(".csv"):
                csv_to_xlsx(inputFileName,outputFileName)
                return {'Status' : "Success"}
            elif inputFileName.endswith(".xlsb"):
                xlsb_to_xlsx(inputFileName,outputFileName)
                return {'Status' : "Success"}
            
            elif (inputFileName.endswith(".xls") or inputFileName.endswith(".xlsm")) or (inputFileName.endswith(".xlt") or inputFileName.endswith(".xltm")):
                xls_to_xlsx(inputFileName,outputFileName)
                return {'Status' : "Success"}
                
            else:
                return {'Exception' : "Non Supported Format found"}

        except Exception as e:
            return {'Exception' : str(e)}

if __name__ == '__main__':
    context = {}
    bot_obj = ConvertDifferentExcelFormatsToXlsx()

    # --input parameters--
    context = {'inputFileName': "",
            'outputFileName': ""}
#    context = {'inputFileName': "Normal.csv",
#            'outputFileName': "Normal_out3.xlsx"}

    resp = bot_obj.execute(context)
    print('response : ',resp)


